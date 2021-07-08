#!/usr/bin/env python
# -*- coding: utf-8 -*-

import time
import datetime
import sys
import argparse
import logging
from openpyxl import load_workbook
import os
import revision
from collections import OrderedDict
import itertools

FRAME_HEADER = "AA"
FRAME_FOOTER = "55"
FRAME_FCT_TAG = "FCT:"
FRAME_DEFAULT_SEP = "-"
FCT_SEP = "_"
MAX_NB_PARAM = 21
MAX_NB_TEST_NAMES = 10


def parse_excel_sheet(book, sheet_name, parse_by='row', max_item=None):
    if parse_by == 'row':
        book = load_workbook(book, read_only=True)
    else:
        book = load_workbook(book)
    d_list = []
    keys = []
    sheet = book.get_sheet_by_name(sheet_name)
    if parse_by == 'row':
        row_index = 0
        for row in sheet.rows:
            fct_dict = {}
            col_index = 0
            for cell in row:
                if row_index:
                    fct_dict[keys[col_index]] = cell.value
                else:
                    keys.append(cell.value)
                col_index += 1
            if row_index:
                d_list.append(fct_dict)

            if max_item is not None:
                if row_index >= max_item:
                    break
            row_index += 1
    elif parse_by == 'col':
        col_index = 0
        for col in sheet.columns:
            fct_dict = {}
            row_index = 0
            for cell in col:
                if col_index:
                    fct_dict[keys[row_index]] = cell.value
                else:
                    keys.append(cell.value)
                row_index += 1
            if col_index:
                d_list.append(fct_dict)
            if max_item is not None:
                if col_index >= max_item:
                    break
            col_index += 1
    return d_list


class ArbSrcGenerator(object):
    def __init__(self, xlfile, xlsheet, rev, fct_prefix, fct_lower, fct_camelcase, fct_sep) -> None:
        super().__init__()
        self.log = logging.getLogger(__name__)
        self.rev = rev
        test_list = parse_excel_sheet(xlfile, xlsheet)

        self.fct_dict = OrderedDict()
        for test in test_list:
            fct_name = fct_prefix
            fct_prototype = "void "
            for i in range(MAX_NB_TEST_NAMES):
                if 'test_name' + str(i) in test:
                    if test['test_name' + str(i)] is not None and test['test_name' + str(i)] != "":
                        fct_name += fct_sep
                        if fct_camelcase:
                            fct_name += test['test_name' + str(i)].title()
                        else:
                            fct_name += test['test_name' + str(i)]

            if fct_lower:
                fct_name = fct_name.lower()
            if fct_name not in self.fct_dict:
                self.fct_dict[fct_name] = {}
                self.fct_dict[fct_name]['test'] = []
                self.fct_dict[fct_name]['param_type_list'] = []
                self.fct_dict[fct_name]['param_values_list'] = []
            else:
                pass

            param_type_list = []
            param_value_list = []

            fct_prototype += fct_name+"("
            for i in range(MAX_NB_PARAM):
                if 'dyn_param' + str(i) in test:
                    if test['dyn_param' + str(i)] is not None and test['dyn_param' + str(i)] != "":
                        if i > 0:
                            fct_prototype += ', '
                        param_list = test['dyn_param' + str(i)].split(';')
                        p_type = param_list[1]
                        p_name = param_list[0]
                        param_value_list.append(eval(param_list[2]))
                        fct_prototype += p_type
                        fct_prototype += ' '
                        fct_prototype += p_name
                        param_type_list.append(p_type)
            fct_prototype += ");"
            self.fct_dict[fct_name]['param_type_list'] = param_type_list
            self.fct_dict[fct_name]['fct_prototype'] = fct_prototype
            self.fct_dict[fct_name]['param_values_list'].append(
                param_value_list)
            self.fct_dict[fct_name]['param_name_list'] = p_name
            if test['comment']:
                comment_list = test['comment'].split('\n')

    def generate_header(self, headerfile):
        outfile = open(headerfile, 'w')
        outfile.write(
            "//Auto-generated file with arb_src_generator {}\n".format(self.rev))
        f_name = os.path.basename(headerfile).upper().replace(".", "_")
        outfile.write("//Automatically  Generated define from config file\n")
        outfile.write("\n#ifndef {}_INCLUDED\n".format(f_name))
        outfile.write("#define {}_INCLUDED\n\n".format(f_name))
        for fct in self.fct_dict:
            outfile.write(self.fct_dict[fct]['fct_prototype'])
            outfile.write("\n")
        outfile.write("\n")
        outfile.write(
            "// Automated Arb_Test Header Generator Builds {0} tests".format(len(self.fct_dict)))
        outfile.write("\n#endif //  {}_INCLUDED\n".format(f_name))
        outfile.close()

    def generate_decoder(self, decoderfile):
        outfile = open(decoderfile, 'w')
        outfile.write(
            "//Auto-generated file with arb_src_generator {}\n".format(self.rev))
        f_name = os.path.basename(decoderfile).upper().replace(".", "_")
        outfile.write("\n#ifndef {}_INCLUDED\n".format(f_name))
        outfile.write("#define {}_INCLUDED\n\n".format(f_name))
        outfile.write("void arb_decoder(const tArbConfig *test)\n{")
        cnt_fct = 0
        for fct in self.fct_dict:
            first_param = True
            cnt_int = 0
            if cnt_fct:
                if_pattern = '\n\t else if'
            else:
                if_pattern = '\n\tif'
            outfile.write(
                "{} (!strcmp({},(const char *)\"{}\"))".format(if_pattern, 'test->fct_name', fct))
            outfile.write("\n\t{{ {}(".format(fct))
            for param in self.fct_dict[fct]['param_type_list']:
                if not first_param:
                    outfile.write(" ,")
                if 'int' in param:
                    outfile.write("test->param_int[{}]".format(cnt_int))
                    cnt_int += 1
                first_param = False
            outfile.write(");}")
            cnt_fct += 1
        outfile.write("\n}")
        outfile.write("\n#endif //  {}_INCLUDED\n".format(f_name))
        outfile.close()

    def generate_jobfile(self, jobfile, sep):
        with open(jobfile, 'w') as outfile:
            outfile.write('\n')
            outfile.write(
                "# Auto-generated file with arb_src_generator {}\n".format(self.rev))
            cnt_fct = 0
            cnt_frames = 0
            for fct in self.fct_dict:
                nb_param_sets = len(self.fct_dict[fct]['param_values_list'])
                outfile.write("# Parameters matrix for {} : {} parameters sets \n".format(
                    fct, nb_param_sets))
                param_set = 0
                for param_vals in self.fct_dict[fct]['param_values_list']:
                    param_set += 1
                    param_matrix = itertools.product(*param_vals)
                    outfile.write("# Parameters set number {}/{} : {} Combinaisons\n".format(
                        param_set, nb_param_sets, len(list(param_matrix))))
                    for current_params in itertools.product(*param_vals):
                        frame = FRAME_HEADER + sep + FRAME_FCT_TAG + fct
                        for v in current_params:
                            frame += sep + str(v)
                        frame += sep + FRAME_FOOTER
                        outfile.write(frame)
                        outfile.write("\n")
                        cnt_frames += 1
                outfile.write("\n")
                outfile.write("# {} frames generated\n".format(cnt_frames))


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--debug', action='store_true')
    parser.add_argument('--debug_console', action='store_true')
    parser.add_argument('--param_file', action='store', default="arb.xlsm")
    parser.add_argument('--test_sheet', action='store', default="tests")
    parser.add_argument('--test_header_file',
                        action='store', default="arb_test.h")
    parser.add_argument('--arb_decoder_file',
                        action='store', default="arb_decoder.h")
    parser.add_argument('--job_file',
                        action='store', default="jobfile.csv")
    parser.add_argument('--serial_sep',
                        action='store', default=FRAME_DEFAULT_SEP)
    parser.add_argument('--fct_prefix',
                        action='store', default="arbtest")
    parser.add_argument('--fct_lower',
                        action='store_true', default=True)
    parser.add_argument('--fct_camelcase',
                        action='store_true', default=False)
    parser.add_argument('--fct_sep',
                        action='store', default=FCT_SEP)

    args = parser.parse_args()

    if args.debug or args.debug_console:
        debug_level = logging.DEBUG
    else:
        debug_level = logging.ERROR
    FORMAT = '%(asctime)s :: %(levelname)s :: %(name)s :: %(funcName)s :: %(message)s'
    if args.debug_console:
        logging.basicConfig(level=debug_level,
                            format=FORMAT, stream=sys.stdout)
    else:
        logging.basicConfig(level=debug_level, format=FORMAT)
    log = logging.getLogger(__name__)
    rev = revision.REVISION + '.' + str(revision.BUILD)
    log.info("Starting {}".format(rev))

    generator = ArbSrcGenerator(args.param_file, args.test_sheet, rev,
                                args.fct_prefix, args.fct_lower, args.fct_camelcase, args.fct_sep)
    generator.generate_header(args.test_header_file)
    generator.generate_decoder(args.arb_decoder_file)
    generator.generate_jobfile(args.job_file, args.serial_sep)

#!/usr/bin on
# -*- coding: utf-8 -*-


import datetime
import time
import sys
from openpyxl import load_workbook
from termcolor import colored

def print_red(msg):
	print(colored(msg,'red'))

def print_green(msg):
	print(colored(msg,'green'))

def print_yellow(msg):
	print(colored(msg,'yellow'))

def get_timestamp():
	"""
	Return a formated string of date+time
	"""
	return datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')

def fmt_eng(num, suffix='', nb_dig=3):
	fmt_big_normal = "%3.{0}f%s%s".format(nb_dig)
	fmt_big_ovf = "%.{0}f%s%s".format(nb_dig)
	fmt_low_normal = "%.{0}f%s%s".format(nb_dig)
	fmt_low_ovf = "%3.{0}f%s%s".format(nb_dig)
	if abs(num)>1:
		for unit in ['','K','M','G','T','P','E','Z']:
			if abs(num) < 1000.0:
				return (fmt_big_normal % (num, unit, suffix)).strip()
			num /= 1000.0
		return (fmt_big_ovf % (num, 'Y', suffix)).strip()

	else:
		num *= 1000.0
		for unit in ['m','µ','n','p','f']:
			if abs(num) > 1.0:
				return (fmt_low_normal % (num, unit, suffix)).strip()
			num *= 1000.0
		return (fmt_low_ovf % (num, 'a', suffix)).strip()


def get_excel_sheet_names(book):
	book = load_workbook(book, read_only=True)
	l = []
	for s in book.get_sheet_names():
		if s != "config":
			l.append(s)
	return l


def parse_excel_sheet(book, sheet_name, parse_by='row', max_item=None):
	if parse_by == 'row':
		book = load_workbook(book, read_only=True, data_only=True)
	else:
		book = load_workbook(book)
	d_list = []
	keys = []
	sheet = book.get_sheet_by_name(sheet_name)
	if parse_by == 'row':
		row_index = 0
		for row in sheet.rows:
			d = {}
			col_index = 0
			for cell in row:
				if row_index:
					d[keys[col_index]] = cell.value
				else:
					keys.append(cell.value)
				col_index += 1
			if row_index:
				d_list.append(d)

			if max_item is not None:
				if row_index >= max_item:
					break;
			row_index += 1
	elif parse_by == 'col':
		col_index = 0
		for col in sheet.columns:
			d = {}
			row_index = 0
			for cell in col:
				if col_index:
					d[keys[row_index]] = cell.value
				else:
					keys.append(cell.value)
				row_index += 1
			if col_index:
				d_list.append(d)
			if max_item is not None:
				if col_index >= max_item:
					break;
			col_index += 1
	del sheet
	del book
	return d_list

def my_excepthook(type, value, tback):
    # log the exception here

    # then call the default handler
    sys.__excepthook__(type, value, tback)

def RANGE(start, stop, step):
    """
Similar as regular range but for float
:param start:
:param stop:
:param step:
"""
    r = []
    step = abs(step)
    v = start
    if stop > start:
	    while v <= stop:
		    r.append(v)
		    v = round(v + step, 3)
    else:
	    while v >= stop:
		    r.append(v)
		    v = round(v - step, 3)
    return r